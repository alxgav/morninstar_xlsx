from openpyxl import load_workbook

from config import path

sheets = ('Enquiry #1', 'Enquiry #2', 'Enquiry #3', 'Enquiry #4', 'Enquiry #5', 'Enquiry #6')

def row_value(item, target, row, col=8):
    for item_row in item:
        value = item_row
        if value is not None:
            target.cell(row, col).value = float(value)
            target.cell(row, col).number_format = '#,##0.00'
        else:
            target.cell(row, col).value = ''
        col +=1

def row_companies(target, item, row):
    target.cell(row, 1).value = item['My CODE']
    target.cell(row, 2).value = item['Symbol']
    target.cell(row, 3).value = item['Exchange']
    target.cell(row, 4).value = item['Company']


def make_xlsx(data_xlsx):
    wb = load_workbook(filename = f'{path}/templates/templates.xlsx')
    ws = wb.active
    row = 2
    target_1 = wb.copy_worksheet(wb['template#1'])
    target_1.title = sheets[0]
    
    ''' Enquiry#1 '''

    for item in data_xlsx:
        row_companies(target_1, item, row)
        if len(item['Enquiry#1']) > 0:
            target_1.cell(row, 5).value = item['Enquiry#1']['Year_end_month']
            target_1.cell(row, 6).value = item['Enquiry#1']['Year_end']
            target_1.cell(row, 7).value = item['Enquiry#1']['currency']
            row_value(item['Enquiry#1']['Total_revenue'], target_1, row, col=8)
            row_value(item['Enquiry#1']['Diluted_net'], target_1, row, col=13)
            row_value(item['Enquiry#1']['Diluted_weighted'], target_1, row, col=18)
            row_value(item['Enquiry#1']['Diluted_ESP'], target_1, row, col=23)
        row +=1
    
    
    ''' Enquiry#2 '''
    
    
    row = 2
    target_2 = wb.copy_worksheet(wb['template#2'])
    target_2.title = sheets[1]
    for item in data_xlsx:
        row_companies(target_2, item, row)
        if len(item['Enquiry#2']) > 0:
            row_value(item['Enquiry#2']['Total_assest'], target_2, row, col=5)
            row_value(item['Enquiry#2']['Total_equity'], target_2, row, col=10)
        row +=1

    
    ''' Enquiry#3 '''
    
    
    row = 2
    target_3 = wb.copy_worksheet(wb['template#3'])
    target_3.title = sheets[2]
    for item in data_xlsx:
        row_companies(target_3, item, row)
        if len(item['Enquiry#3']) > 0:
            row_value(item['Enquiry#3']['Equity'], target_3, row, col=5)
            row_value(item['Enquiry#3']['Invested_capital'], target_3, row, col=15)
        row +=1
    

    ''' Enquiry#4 '''
    
    
    row = 2
    target_4 = wb.copy_worksheet(wb['template#4'])
    target_4.title = sheets[3]
    for item in data_xlsx:
        row_companies(target_4, item, row)
        if len(item['Enquiry#4']) > 0:
            row_value(item['Enquiry#4']['Current_ratio'], target_4, row, col=5)
            row_value(item['Enquiry#4']['Book_value'], target_4, row, col=15)
        row +=1

    ''' Enquiry#5 '''
    
    row = 2
    target_5 = wb.copy_worksheet(wb['template#5'])
    target_5.title = sheets[4]
    for item in data_xlsx:
        row_companies(target_5, item, row)
        if len(item['Enquiry#5']) > 0:
            row_value(item['Enquiry#5']['capExAsPerOfSales'], target_5, row, col=5)
            row_value(item['Enquiry#5']['freeCashFlowPerShare'], target_5, row, col=15)
        row +=1
    

    ''' Enquiry#6 '''
    
    row = 2
    target_6 = wb.copy_worksheet(wb['template#6'])
    target_6.title = sheets[5]
    for item in data_xlsx:
        row_companies(target_6, item, row)
        if len(item['Enquiry#6']) > 0:
            row_value(item['Enquiry#6']['Dividend_Per_Share'], target_6, row, col=5)
        row +=1

    wb.remove(wb['template#1'])  
    wb.remove(wb['template#2'])
    wb.remove(wb['template#3']) 
    wb.remove(wb['template#4'])
    wb.remove(wb['template#5'])
    wb.remove(wb['template#6'])
    wb.save(f'{path}/out/morningstar.xlsx')
